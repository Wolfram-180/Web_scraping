import time
from selenium import webdriver
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import datetime
import re
from openpyxl import load_workbook

debug = True
WebDriverWaitInSec = 30
binary = FirefoxBinary('c:\geckodriver\geckodriver.exe')

num_=0

def clean(txt):
    if txt == None:
        txt = ""
    txt = str.strip(txt)
    txt = str.replace(txt, "\"\"", "")
    txt = str.replace(txt, "\"", "")
    txt = str.replace(txt, "'", "")
    txt = txt.replace('\n', ' ')
    txt = txt.replace('\r', '')
    return txt

def init_driver():
    binary = r'c:\Program Files (x86)\Mozilla Firefox\firefox.exe'
    options = Options()
    options.binary = binary
    cap = DesiredCapabilities().FIREFOX
    cap["marionette"] = True #optional
    return webdriver.Firefox(options=options, capabilities=cap, executable_path="c:\\geckodriver\\geckodriver.exe")


def sleep(secs, place='whatever'):
    ttlsecs = secs
    while secs > 0:
        time.sleep(1)
        print('now: {} - {} of {} in {}'.format(datetime.datetime.now(), secs, ttlsecs, place))
        secs -= 1

additional_search_word = ' мазут'
additional_reqs = ['']
exclusing_word = ['карт', 'карты']
period = '01.01.2019-31.12.2019'  # зашит в начальную ссылку

xlsx_end_row = 500
xlsx_links_row = 1

if __name__ == "__main__":
    wb = load_workbook(filename='data.xlsx')
    ws = wb['data']
    wslinks = wb['links']

    total_orders_count = 0
    order_view_browser = init_driver()
    search_mainword = init_driver()
    sleep(10, 'start')
    # открываем страницу Закупки
    search_mainword.get('https://zakupki.gov.ru/epz/order/extendedsearch/results.html?morphology=on&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F&pageNumber=1&sortDirection=false&recordsPerPage=_10&showLotsInfoHidden=false&fz44=on&fz223=on&sortBy=UPDATE_DATE&af=on&ca=on&pc=on&pa=on&publishDateFrom=01.01.2019+-+31.12.2019&publishDateTo=31.12.2019&currencyIdGeneral=-1')
    sleep(10, 'search_mainword.get')
    try:
        body = search_mainword.find_element_by_css_selector('body')
   
        # находим и кликаем на выпадающий список Показать по
        isSet50 = False
        if isSet50:
            field50 = search_mainword.find_element_by_xpath('//*[@id="quickSearchForm_header"]/section[2]/div/div/div[1]/div[4]/div/div[2]/div[2]/div[1]/div')
            field50.click()
            sleep(5, 'field50.dropdown')
            # выбираем 50
            field50 = search_mainword.find_element_by_xpath('//*[@id="_50"]')
            field50.click()
            sleep(5, 'field50.click on 50')

        print(ws.dimensions)
        print("Minimum row: {0}".format(ws.min_row))
        print("Maximum row: {0}".format(ws.max_row))
        print("Minimum column: {0}".format(ws.min_column))
        print("Maximum column: {0}".format(ws.max_column))

        for r in range(1, ws.max_row + 1):
            if r > xlsx_end_row:
                pass
            else:
                is_todo = ws['A'+str(r)].value
                if is_todo == 'x':
                    isNextBtn = True
                    search_word = ws['C'+str(r)].value + additional_search_word  # наименование компании + топливо
                    ttlcost = 0
                    # вводим в поле поиска искомую фразу
                    searchfield = search_mainword.find_element_by_xpath('//*[@id="searchString"]')
                    searchfield.clear()
                    searchfield.send_keys(search_word)
                    sleep(1, 'searchfield.send_keys(search_word)')
                    # кликаем лупу
                    searchbtn = search_mainword.find_element_by_xpath('//*[@id="quickSearchForm_header"]/section[1]/div/div/div/div[2]/div/div/button')
                    searchbtn.click()
                    sleep(10, 'searchbtn.click()')

                    while isNextBtn == True:
                        # нажимаем 20 раз PGDOWN PGUP на основной странице для прогрузки
                        body = search_mainword.find_element_by_css_selector('body')
                        for i in range(1, 20):
                            body.send_keys(Keys.PAGE_DOWN)
                            time.sleep(0.1)
                        for i in range(1, 20):
                            body.send_keys(Keys.PAGE_UP)
                            time.sleep(0.1)

                        # получаем список 1я страница
                        links = search_mainword.find_elements_by_class_name('registry-entry__header-mid__number')
                        orderlinksearch_list = []
                        orderscost_list = []
                        for link in links:
                            linktext = link.find_element_by_tag_name("a").get_attribute('href')
                            try:
                                cost_text = link.find_element_by_xpath('//*[@id="quickSearchForm_header"]/section[2]/div/div/div[1]/div[3]/div[1]/div/div[2]/div[1]/div[2]').text
                            except:
                                cost_text = '0'
                            cost_text = cost_text.replace(' ', '')
                            cost_text = cost_text.replace(',', '.')
                            cost_text = cost_text.replace('₽', '')

                            cost = float(cost_text)
                            orderlinksearch_list.append(linktext)
                            orderscost_list.append(cost)
                            sleep(1, linktext)

                        pos = 0
                        for orderlinksearch in orderlinksearch_list:
                            print('pos: {}'.format(pos))
                            order_view_browser.get(orderlinksearch)
                            sleep(5, 'order_view_browser.get(orderlinksearch)')
                            # нажимаем 5 раз PGDOWN PGUP на странице закупки для прогрузки
                            body = order_view_browser.find_element_by_css_selector('body')
                            SectionAddDataFoundAndClicked = False
                            SectionAddData = None
                            for i in range(1, 5):
                                body.send_keys(Keys.PAGE_DOWN)
                                time.sleep(0.25)
                                # открываем секцию Требования, если есть
                                if not SectionAddDataFoundAndClicked:
                                    try:
                                        SectionAddData = order_view_browser.find_element_by_xpath('/html/body/div[4]/div/div[9]/div/div/div/div[1]/span[2]').click()
                                        SectionAddDataFoundAndClicked = True
                                    except Exception as ex:
                                        SectionAddDataFoundAndClicked = False
                            for i in range(1, 5):
                                body.send_keys(Keys.PAGE_UP)
                                time.sleep(0.25)
                            # ищем слова в тексте всей страницы
                            exclusing_word_found = []
                            src = order_view_browser.page_source
                            for excl_ in exclusing_word:
                                isExclFound = None
                                try:
                                    isExclFound = re.search(r'{}'.format(excl_), src)
                                    sleep(2, 're.search')
                                except Exception as ex:
                                    isExclFound = None
                                if isExclFound is not None:
                                    exclusing_word_found.append(excl_)

                            # не нашли слово-исключение - суммируем кост в общее значение по компании
                            if len(exclusing_word_found) == 0:
                                ttlcost = ttlcost + orderscost_list[pos]
                                emailtext = orderlinksearch + ' \n\n ' + str(orderscost_list[pos])

                                xlsx_links_row += 1
                                wslinks['A' + str(xlsx_links_row)].value = search_word
                                wslinks['B' + str(xlsx_links_row)].value = orderlinksearch
                                wslinks['C' + str(xlsx_links_row)].value = str(orderscost_list[pos])

                            pos += 1
                        try:
                            btnnext = search_mainword.find_element_by_css_selector('#quickSearchForm_header > section.content.content-search-registry-block > div > div > div.col-9.search-results > div.paginator-block > div > div.paginator.align-self-center.m-0 > ul > a.paginator-button.paginator-button-next')
                            btnnext.click()
                        except Exception as ex:
                            isNextBtn = False

                    ws['F' + str(r)].value = ttlcost
                    ws['B' + str(r)].value = 'x'
                    wb.save('data.xlsx')

    finally:
        search_mainword.quit()
        order_view_browser.quit()
        wb.save('data.xlsx')